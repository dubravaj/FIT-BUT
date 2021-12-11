import sys
import os
import re
import json
import openpyxl

from Bio import pairwise2

from fireprotdb.importer.db import *

from fireprotdb.importer.utils import cache_file, data_file
from fireprotdb.importer.utils.alignment import align
from fireprotdb.importer.utils.pdb import load_pdb, fetch_pdb_metadata
from fireprotdb.importer.utils.pubmed import map_pubmed_ids, explodepub
from fireprotdb.importer.utils.uniprot import fetch_uniprot_entry, UniprotObsoleteMapping
from fireprotdb.importer.protherm.converter import uniprot_mappings as protherm_uniprot_mappings

wb = openpyxl.load_workbook(data_file("literature_search/dataset2.xlsx"), data_only=True)
ws = wb['Singlepoint']

COL_ID = 0
COL_PROTEIN = 1
COL_GENBANK = 2
COL_UNIPROT = 3
COL_PDB = 4
COL_PDB_CHAIN = 5
COL_PDB_MUT = 6
COL_MUT = 7
COL_SOURCE = 8
COL_IS_DERIVED = 9
COL_DERIVED_TYPE = 10
COL_DDG = 11
COL_DTM = 12
COL_TM = 13
COL_DT50 = 14
COL_T50 = 15
COL_REF = 16
COL_PMID = 17
COL_DOI = 18
COL_HALF_LIFE = 19
COL_CM = 20
COL_M = 21
COL_SCAN_RATE = 22
COL_PH = 23
COL_METHOD = 24
COL_DENATURANT = 25
COL_ACTIVE_SITE = 26

uniprot_mappings = UniprotObsoleteMapping({
    ("P51698", "1MJ5"): "D4Z2G1",
})
uniprot_mappings.data.update(protherm_uniprot_mappings.data)

UNMAPPABLE_PUBS = {
    "10.1016/j.compchemeng.2018.05.014": {
        "doi": "10.1016/j.compchemeng.2018.05.014",
        "title": "Computational design of thermostable mutants for cephalosporin C acylase from Pseudomonas strain SE83",
        "journal": "Computers & Chemical Engineering",
        "volume": "116",
        "issue": None,
        "year": "2018",
        "pages": "112-21",
        "authors": [
            {
                "last_name": "Xu",
                "fore_name": "Zhaobin",
                "initials": "Z"
            },
            {
                "last_name": "Tian",
                "fore_name": "Ye",
                "initials": "Y"
            },
            {
                "last_name": "Zhu",
                "fore_name": "Yushan",
                "initials": "Y"
            }
        ],
    },
    "10.1021/acscatal.9b05223": {
        "doi": "10.1021/acscatal.9b05223",
        "title": "Robust ω-Transaminases by Computational Stabilization of the Subunit Interface",
        "journal": "ACS Catalysis",
        "volume": "10",
        "issue": "5",
        "year": "2020",
        "pages": "2915-28",
        "authors": [
            {
                "last_name": "Meng",
                "fore_name": "Qinglong",
                "initials": "Q"
            },
            {
                "last_name": "Capra",
                "fore_name": "Nikolas",
                "initials": "N"
            },
            {
                "last_name": "Palacio",
                "fore_name": "Cyntia M.",
                "initials": "CM"
            },
            {
                "last_name": "Lanfranchi",
                "fore_name": "Elisa",
                "initials": "E"
            },
            {
                "last_name": "Otzen",
                "fore_name": "Marleen",
                "initials": "M"
            },
            {
                "last_name": "van Schie",
                "fore_name": "Luc Z.",
                "initials": "LZ"
            },
            {
                "last_name": "Rozeboom",
                "fore_name": "Henriette J.",
                "initials": "HJ"
            },
            {
                "last_name": "Thunnissen",
                "fore_name": "Andy-Mark W. H.",
                "initials": "AMWH"
            },
            {
                "last_name": "Wijma",
                "fore_name": "Hein J.",
                "initials": "HJ"
            },
            {
                "last_name": "Janssen",
                "fore_name": "Dick B.",
                "initials": "DB"
            },
        ],
    },
    "10.1021/acscatal.6b01062": {
        "doi": "10.1021/acscatal.6b01062",
        "title": "Versatile Peptide C-Terminal Functionalization via a Computationally Engineered Peptide Amidase",
        "journal": "ACS Catalysis",
        "volume": "6",
        "issue": "8",
        "year": "2016",
        "pages": "5405-414",
        "authors": [
            {
                "last_name": "Wu",
                "fore_name": "Bian",
                "initials": "B"
            },
            {
                "last_name": "Wijma",
                "fore_name": "Hein J.",
                "initials": "HJ"
            },
            {
                "last_name": "Song",
                "fore_name": "Lu",
                "initials": "L"
            },
            {
                "last_name": "Rozeboom",
                "fore_name": "Henriette J.",
                "initials": "HJ"
            },
            {
                "last_name": "Poloni",
                "fore_name": "Claudia",
                "initials": "C"
            },
            {
                "last_name": "Tian",
                "fore_name": "Yue",
                "initials": "Y"
            },
            {
                "last_name": "Arif",
                "fore_name": "Muhammad I.",
                "initials": "MI"
            },
            {
                "last_name": "Nuijens",
                "fore_name": "Timo",
                "initials": "T"
            },
            {
                "last_name": "Quaedflieg",
                "fore_name": "Peter J. L. M.",
                "initials": "PJLM"
            },
            {
                "last_name": "Szymanski",
                "fore_name": "Wiktor",
                "initials": "W"
            },
            {
                "last_name": "Feringa",
                "fore_name": "Ben L.",
                "initials": "BL"
            },
            {
                "last_name": "Janssen",
                "fore_name": "Dick B.",
                "initials": "DB"
            }
        ],
    },
}

row_range = range(2, ws.max_row + 1)

session = Session()

pubs = cache_file("litsearch_pubs.json")
if not os.path.isfile(pubs):
    with open(pubs, 'w') as outfile:
        pmids = set([ws[i][COL_PMID].value for i in row_range if ws[i][COL_PMID].value])
        json.dump(map_pubmed_ids(list(pmids)), outfile, indent=4)

with open(pubs, 'r') as fh:
    pubs_data = json.load(fh)
    pubs_data.update(UNMAPPABLE_PUBS)

pubs = {}
for (pmid, doi, title, journal, volume, issue, year, pages, authors, _) in [explodepub('pub', p) + (p['authors'], 1) for p in pubs_data.values()]:
    db_pub = session.query(Publication).filter(Publication.doi == doi).filter(Publication.pmid == pmid).first()
    if not db_pub:
        db_pub = Publication(doi=doi, pmid=pmid, title=title, journal=journal, volume=volume, issue=issue, year=year, pages=pages)
        session.add(db_pub)
        session.commit()
        author_order = 1
        for a in authors:
            (last_name, fore_name, initials) = (a['last_name'], a['fore_name'], a['initials'])
            a = Author(last_name=last_name, fore_name=fore_name, initials=initials)
            session.add(a)
            session.commit()
            session.add(AuthorsPublications(publication_id=db_pub.publication_id, author_id=a.author_id, author_order=author_order))
            author_order += 1
    pubs[(pmid, doi.lower())] = db_pub

session.commit()

TO_SKIP = [
    # singlepoint mutants of multiple-point mutant
    'AGT-LM'
]

not_imported = 0
errors = []
myErrors = list()

for i in row_range:
    row = ws[i]
    values = [ c.value if c else None for c in row ]
    print("Processing row {}".format(values))

    experiment_id = values[COL_ID]

    if experiment_id == None:
        break

    protein = values[COL_PROTEIN]

    if protein in TO_SKIP:
        not_imported += 1
        continue

    pdb_id = values[COL_PDB]
    uniprot_ac = uniprot_mappings.map(values[COL_UNIPROT], pdb_id)

    m = re.match("([A-Z])(\\d+)([A-Z])", values[COL_MUT])
    if not m:
        errors.append((values, "Unable to parse mutation"))
        not_imported += 1
        continue

    wt = m.group(1)
    seq_pos = int(m.group(2))
    mut = m.group(3)

    # we have no sequence information
    if not uniprot_ac and not pdb_id:
        errors.append((values, "We do not have either UniProt nor PDB code"))
        not_imported += 1
        continue

    if uniprot_ac:
        try:
            uniprot = fetch_uniprot_entry(uniprot_ac)
        except Exception as e:
            myErrors.append([values[0], values[8], "unable to catch uniprot"])
            continue
        seq = uniprot['id_sequence']

        db_seq = session.query(ProteinSequence).filter(ProteinSequence.protein_seq == seq).first()
        if db_seq:
            db_unp = session.query(Uniprot).filter(Uniprot.uniprot_id == uniprot_ac).first()
            if not db_unp:
                db_unp = Uniprot(uniprot_id=uniprot_ac, sequence_id=db_seq.sequence_id, relation='alternative')
                session.add(db_unp)
                session.commit()
        else:
            db_seq = ProteinSequence(protein_seq=seq, protein_name=uniprot['id_name'], uniprot_id=uniprot_ac, species=uniprot['id_organism'], ec_number=uniprot['ec_number'])
            session.add(db_seq)
            session.commit()

        for my_pdb_id, pdb_entry in uniprot['pdb_mapping'].items():
            db_pdb = session.query(Structure).filter(Structure.pdb_id == my_pdb_id).first()
            if not db_pdb:
                pdb_method = ""
                pdb_resolution = ""
                try:
                    pdb_method = method=pdb_entry['method']
                    pdb_resolution = pdb_entry['resolution']
                except:
                    None
                db_pdb = Structure(pdb_id=my_pdb_id, method=pdb_method, resolution=pdb_resolution)
                session.add(db_pdb)
                session.commit()

            for pdb_chain in pdb_entry['chains'].keys():
                db_pdb_mapping = session.query(StructureSequence)\
                    .filter(StructureSequence.pdb_id == my_pdb_id) \
                    .filter(StructureSequence.chain == pdb_chain) \
                    .filter(StructureSequence.sequence_id == db_seq.sequence_id)\
                    .first()
                if not db_pdb_mapping:
                    db_pdb_mapping = StructureSequence(pdb_id=my_pdb_id, chain=pdb_chain, sequence_id=db_seq.sequence_id)
                    session.add(db_pdb_mapping)

            session.commit()

    if pdb_id:
        try:
            seqs, uniprot_data, struct_index_mapping, missing_residues, seqres = load_pdb(pdb_id, uniprot_mappings)
        except Exception as e:
            myErrors.append([values[0], values[8], "unable to parse pdb"])
            continue

        db_pdb = session.query(Structure).filter(Structure.pdb_id == pdb_id).first()
        if not db_pdb:
            pdb_meta = fetch_pdb_metadata(pdb_id)
            pdb_method = ""
            pdb_resolution = ""
            try:
                pdb_method = method = pdb_entry['method']
                pdb_resolution = pdb_entry['resolution']
            except:
                None
            db_pdb = Structure(pdb_id=pdb_id, method=pdb_method, resolution=pdb_resolution)
            session.add(db_pdb)
            session.commit()

        #TODO: what if PDB is not in mapping?
        db_bio_unit = session.query(BiologicalUnit).filter(BiologicalUnit.pdb_id == pdb_id).first()
        if not db_bio_unit:
            db_bio_unit = BiologicalUnit(pdb_id=pdb_id, pdb_unit_number=1)
            session.add(db_bio_unit)
            session.commit()

        if not uniprot_ac:
            chain = values[COL_PDB_CHAIN]
            if not chain:
                if len(seqs) > 1:
                    errors.append((values, "No chain specified for multimer PDB without Uniprot AC"))
                    not_imported += 1
                    continue
                chain = next(iter(seqres.keys()))

            seq = seqres[chain]

            db_seq = session.query(ProteinSequence).filter(ProteinSequence.protein_seq == seq).first()
            if not db_seq:
                db_seq = ProteinSequence(protein_seq=seq, protein_name=protein)
                session.add(db_seq)
                session.commit()
        else:
            if pdb_id in uniprot['pdb_mapping']:
                #TODO: multiple chains for one entry
                chain = next(iter(uniprot['pdb_mapping'][pdb_id]["chains"].keys()))
                print("Selecting chain {}".format(chain))
            else:
                errors.append((values, "Chain was not found for uniprot entry {}".format(uniprot_ac)))
                not_imported += 1
                continue

        db_seq_biounit = session.query(ProteinSequenceBiologicalUnit) \
            .filter(ProteinSequenceBiologicalUnit.sequence_id == db_seq.sequence_id) \
            .filter(ProteinSequenceBiologicalUnit.bio_unit_id == db_bio_unit.bio_unit_id)\
            .filter(ProteinSequenceBiologicalUnit.old_chain == chain).first()
        if not db_seq_biounit:
            db_seq_biounit = ProteinSequenceBiologicalUnit(sequence_id=db_seq.sequence_id, bio_unit_id=db_bio_unit.bio_unit_id, old_chain=chain, new_chain=chain)
            session.add(db_seq_biounit)
            session.commit()

        pdb_seq = seqs[chain]
        aln = align(pdb_seq, seq)

        if len(seq) < seq_pos - 1:
            myErrors.append([values[0], values[8], "seq len error"])
            continue

        if seq[seq_pos - 1] != wt or not aln.two_one(seq_pos):
            if seq_pos <= len(pdb_seq) and pdb_seq[seq_pos - 1] == wt:
                pdb_seq_pos = seq_pos
                seq_pos = aln.one_two(seq_pos)
            elif seq_pos in struct_index_mapping[chain] and pdb_seq[struct_index_mapping[chain][seq_pos] - 1] == wt:
                pdb_seq_pos = struct_index_mapping[chain][seq_pos]
                seq_pos = aln.one_two(pdb_seq_pos)
            else:
                errors.append((values, "Mismatch residue for entry {} {}: {} {} {}".format(uniprot_ac, pdb_id, seq[seq_pos - 1], seq_pos, wt)))
                not_imported += 1
                continue
        else:
            pdb_seq_pos = aln.two_one(seq_pos)
            if not pdb_seq_pos:
                errors.append((values, "cannot map seq pos {} {}\n{}\n{}".format(seq_pos, wt, aln.seqA, aln.seqB)))
                not_imported += 1
                continue
            for (k, v) in struct_index_mapping[chain].items():
                if v == pdb_seq_pos:
                    struct_pos = k
                    print("Structure index in PDB is {}".format(struct_pos))
                    break
    else:
        if seq[seq_pos - 1] != wt:
            errors.append((values, "Mismatch residue for entry {} {}: {} {} {}".format(uniprot_ac, pdb_id, seq[seq_pos - 1], seq_pos, wt)))
            not_imported += 1
            continue

    if values[COL_TM] == "-":
        values[COL_TM] = None
    if values[COL_DTM] == "-":
        values[COL_DTM] = None
    if values[COL_DDG] == "-":
        values[COL_DDG] = None
    ddg = float(values[COL_DDG]) if values[COL_DDG] is not None else None
    dtm = float(values[COL_DTM]) if values[COL_DTM] is not None else None
    tm = float(values[COL_TM]) if values[COL_TM] is not None else None

    db_res = session.query(Residue).filter(Residue.sequence_id == db_seq.sequence_id).filter(Residue.position == seq_pos).first()
    if not db_res:
        try:
            db_res = Residue(sequence_id=db_seq.sequence_id, position=seq_pos, residue=wt)
            session.add(db_res)
            session.commit()
        except Exception as e:
            myErrors.append([values[0], values[8], "unable to map residue"])
            session.rollback()
            continue

    if pdb_id:
        db_struct_res = session.query(StructureResidue).filter(StructureResidue.sequence_id == db_seq.sequence_id).filter(StructureResidue.bio_unit_id == db_bio_unit.bio_unit_id).filter(StructureResidue.position == seq_pos).first()
        if not db_struct_res:
            db_struct_res = StructureResidue(sequence_id=db_seq.sequence_id, bio_unit_id=db_bio_unit.bio_unit_id, position=seq_pos, structure_index=struct_pos, new_chain=chain)
            session.add(db_struct_res)
            session.commit()

    db_mut = session.query(Mutation).filter(Mutation.sequence_id == db_seq.sequence_id).filter(Mutation.position == seq_pos).filter(Mutation.mutated_aa == mut).first()
    if not db_mut:
        db_mut = Mutation(sequence_id=db_seq.sequence_id, position=seq_pos, mutated_aa=mut)
        session.add(db_mut)
        session.commit()

    pmid = str(values[COL_PMID]) if values[COL_PMID] else None
    doi = values[COL_DOI].replace('–', '-').lower() if values[COL_DOI] else None

    db_mutexp = session.query(MutationExperiment).filter(MutationExperiment.experiment_id == experiment_id).first()
    if not db_mutexp:
        db_mutexp = MutationExperiment(
            experiment_id=experiment_id,
            mut_id=db_mut.mut_id,
            publication_id=None if not pmid and not doi else pubs[(pmid, doi)].publication_id,
            ddg=ddg,
            d_tm=dtm,
            tm=tm,
            currated=True,
            pH=float(values[COL_PH]) if values[COL_PH] else None,
            method=values[COL_METHOD],
            method_details=None,
            technique=None,
            technique_details=None,
            notes=None
        )

        session.add(db_mutexp)
        session.commit()

test = open("/home/fireprotasr/fireprotdb/fireprotdb-importer/data/literature_search/errors.txt", "w")
for err in myErrors:
    test.write(str(err) + "\n")

for (v,e) in errors:
    test.write(str(v) + "," + str(e) + "\n")

#print("\n\nERRORS:")
#for (v, e) in errors:
#    print("> {}".format(v))
#    print("\t{}".format(e.replace("\n", "\n\t")))
#print("\n\n-------------\nNOT IMPORTED MUTATIONS: {}".format(not_imported))
#wb.close()

#NOTES:
# - priste urcit v jakem mapovani je index rezidua (takhle je to debilni magie)
# - nejasne nazvy publikaci (Li, 2018 je tam dvakrat)
# - Nikolova, PV 2000 neexistuje
# - Joerger AC, 2005/2006 + Ang HC, 2006 + Kather, 2008 + Raquet, 1995 + Berhstein, 2008 + Dubus, 1994 + Fowler, 2011 + Acevedo, 2017 uplne chybi mezi PDF
# - jsou tam zaznamy kde nejsou zadne informace o sekvenci
# - dalsi sloupce pro info o proteinu
# - u nekterych vubec nesedi cislovani